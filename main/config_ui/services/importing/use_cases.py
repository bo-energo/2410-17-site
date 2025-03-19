from typing import List
from pathlib import Path
from openpyxl import Workbook, load_workbook
from .manager import ImportManager
from .import_schemes import imports_all_data


def __import_by_manager_list(wb: Workbook, manager_list: List[ImportManager]):
    """Импортировать по списку менеджеров."""
    cache = {}
    for manager in manager_list:
        manager.set_cache(cache)
        manager.import_all(wb)


def import_in_migration(localization_file: Path, import_mngs: list[ImportManager]):
    """
    Миграция конфигурации страницы состояния актива
    во время начальных миграций
    """
    __import_by_manager_list(
        wb=load_workbook(localization_file, data_only=True),
        manager_list=import_mngs
    )


def import_all_data(localization_file: Path):
    """Импорт всех данных конфигурации страницы состояния актива из xls"""
    __import_by_manager_list(
        wb=load_workbook(localization_file, data_only=True),
        manager_list=imports_all_data
    )
