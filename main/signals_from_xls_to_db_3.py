import django
import os

from django.db import models
from typing import Dict, List, Union, Iterable
from datetime import datetime
from copy import deepcopy
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet

os.environ['DJANGO_SETTINGS_MODULE'] = 'main.settings'
django.setup()

from dashboard.models import (AccessPoints, Signals, SignalsGuide,
                              SignalСategories, SignalGroups,
                              ModbusTypes,
                              MeasureUnits, Assets, AssetsType, DeviceModels,
                              DeviceTypes, Devices, Protocols, Formulas,
                              DatabusSources, DynamicStorages,
                              Schedules, DataTypes, PlotTypes, SignalTypes,
                              ChartTabs, SignalsChartTabs,
                              Substations, GeoMap, GeoMapSetting,
                              AssetsTypeChartTabs)
from dashboard.utils import guid

# Внимание! Данная версия (v.3) для настройки графиков сигналов
#  в разрезе экземпляров оборудования!!!

# Скрипт заполнения базы информацией о сигналах из xls-файла.
# Создаются сигналы и все связанные экземпляры справочников.
# Внимание!!! Перед выполнением необходимо остановить запущенные локально
# экземпляры приложения (иначе возникает ошибка логирования приложения)!


log_path = "logs\\"
# Полный путь до xls файла с списокм сигналов
signals_file_name = os.getenv("SIGNALS_FILE")
# Полный путь до xls файла с data model
data_models_file_name = os.getenv("DATA_MODEL")
# Полный путь до xls файла с UI model
ui_models_file_name = os.getenv("UI_MODEL")
# sheet_names = ["Сигналы онлайн",]
sgn_list_wb: Workbook = load_workbook(signals_file_name, data_only=True)
d_models_wb: Workbook = load_workbook(data_models_file_name, data_only=True)
ui_models_wb: Workbook = load_workbook(ui_models_file_name, data_only=True)
# Определяется только создаются или сохраняются в БД новые объекты
if (arg := os.getenv("SIGNALS_TO_DB")) and arg.lower() == "true":
    to_db = True
else:
    to_db = False
USING_DB = "default"


class Logger:
    __file_type = "log"
    __sep = "; "
    __record_sep = "\n"
    """Пользовательский класс логирования"""
    def __init__(self, name):
        self.name = name

    def info(self, model: str = None, ws: str = None, *message):
        self.write("INFO!", model, ws, *message)

    def warning(self, model: str = None, ws: str = None, *message):
        self.write("WARNING!", model, ws, *message)

    def error(self, model: str = None, ws: str = None, *message):
        self.write("ERROR!", model, ws, *message)

    def write(self, level: str, *message):
        with open(f"{self.name}.{self.__file_type}", "a", encoding="UTF8") as fp:
            fp.write(f"[{datetime.utcnow().isoformat()}]; {level:<10}; ")
            for line in message:
                fp.write(f"{line}{self.__sep}")
            fp.write(self.__record_sep)

    def operation_info(self, log_name_siffix, *message):
        with open(f"{self.name}_{log_name_siffix}.{self.__file_type}", "a", encoding="UTF8") as fp:
            fp.write(f"[{datetime.utcnow().isoformat()}]; ")
            for line in message:
                fp.write(f"{line}{self.__sep}")
            fp.write(self.__record_sep)

    def end_to_log(self, log_name_siffix, *message):
        file_path = f"{self.name}_{log_name_siffix}.{self.__file_type}"
        if os.path.exists(file_path):
            with open(file_path, "a", encoding="UTF8") as fp:
                fp.write(f"[{datetime.utcnow().isoformat()}] ")
                for line in message:
                    fp.write(f"{line}{self.__sep}")
                fp.write(self.__record_sep)
    #   End class Logger


_LOGGER = Logger(f"{log_path}!filling_db_signals")
_LOGGER.info(None, None, " --- START! ---")


class Manager:
    def __init__(self, db_model: models.Model, local_storage: 'LocalStorage',
                 execution_result: 'Result'):
        self.__db_model = db_model
        self.__local_storage = local_storage
        self.__execution_result = execution_result

    def __create(self, storage: Union[models.Manager, 'LocalStorage'],
                 logger: 'Logger' = None, **kwargs):
        # print(f"{kwargs = }")
        try:
            instance, created = storage.get_or_create(**kwargs)
        except Exception as ex:
            instance = None
            created = False
            if logger:
                err_str = f"[ERROR] при создании {self.__db_model}: {str(ex)}"
                logger.info(err_str)
            print(err_str)
        return instance, created

    def __get(self, storage: Union[models.Manager, 'LocalStorage'],
              logger: 'Logger' = None, **kwargs):
        try:
            instance = storage.get(**kwargs)
        except Exception as ex:
            instance = None
            if logger:
                err_str = f"[ERROR] при поиске {self.__db_model} c условием {kwargs}: {str(ex)}"
                logger.info(err_str)
            print(err_str)
        return instance

    def __update(self, instance: object,
                 logger: 'Logger', **kwargs):
        changes_attrib = {}
        need_to_save = False

        if instance is None:
            return instance, need_to_save, changes_attrib
        copy_instance = deepcopy(instance)
        try:
            for arg, value in kwargs.items():
                value_before = getattr(instance, arg)
                if not value_before == value:
                    need_to_save = True
                    setattr(copy_instance, arg, value)
                    changes_attrib[arg] = {
                        "before": value_before,
                        "after": value
                    }
        except Exception as ex:
            need_to_save = False
            if logger:
                logger.info(str(ex))
            print(ex)
        return copy_instance, need_to_save, changes_attrib

    def __save(self, instance, logger: 'Logger', arg_to_create: dict, to_db: bool):
        if to_db:
            try:
                instance.save(using=USING_DB)
                success = True
            except Exception as ex:
                success = False
                if logger:
                    logger.info(str(ex))
                print(ex)
        else:
            success = self.__local_storage.save(instance, arg_to_create)
        return instance, success

    @staticmethod
    def __remove_xls_pkey(args: dict):
        args.pop("pkey", None)

    def _get_instance(self, args_to_create: dict, args_to_update: dict,
                      row, to_db: bool):
        """Получить сущность класса entity_cls или None"""
        # print(f"{args_to_create = }")
        if to_db:
            storage = self.__db_model.objects.using(USING_DB)
            self.__remove_xls_pkey(args_to_create)
        else:
            storage = self.__local_storage
        # print(f" ===== {self.__db_model}: {storage = }")
        instance, created = self.__create(
            storage,
            logger=self.__execution_result.logger,
            **args_to_create
        )
        # print(f"{instance = }  {created = }  {args_to_update = }")
        if instance and args_to_update:
            updated_instance, need_to_save, changes_attrib = self.__update(
                instance,
                logger=self.__execution_result.logger,
                **args_to_update
            )
        else:
            updated_instance = None
            need_to_save = False
            changes_attrib = {}
        # print(f"{need_to_save = }")
        if need_to_save:
            instance, updated = self.__save(updated_instance,
                                            self.__execution_result.logger,
                                            args_to_create, to_db)
        else:
            updated = None
        # print(f"{updated = }  {instance = }")
        self.__execution_result.write_result(created, updated, obj=instance,
                                             create_kwargs=args_to_create,
                                             update_kwargs=args_to_update,
                                             changes_attrib=changes_attrib,
                                             row=row)
        return instance

    def _get_exist_instance(self, args_to_find: dict,
                            to_db: bool):
        """Получить сущность класса entity_cls или None"""
        # print(f"{args_to_create = }")
        if to_db:
            storage = self.__db_model.objects.using(USING_DB)
            self.__remove_xls_pkey(args_to_find)
        else:
            storage = self.__local_storage
        # print(f" ===== {self.__db_model}: {storage = }")
        instance = self.__get(
            storage,
            logger=self.__execution_result.logger,
            **args_to_find
        )
        return instance

    @staticmethod
    def check_is_dict(args):
        if not isinstance(args, dict):
            args = {}
        return args


class LocalStorage:
    def __init__(self, model: models.Model):
        self.__model = model
        self.__storage = {}

    def __get_key(self, values: dict):
        return values.get("pkey")

    def get(self, **kwargs):
        pkey = self.__get_key(kwargs)
        return self.__storage[pkey]

    def get_or_create(self, **kwargs):
        defaults = kwargs.pop("defaults", None)
        created_key = self.__get_key(kwargs)

        if instance := self.__storage.get(created_key):
            created = False
        else:
            kwargs.pop("pkey", None)
            if isinstance(defaults, dict):
                kwargs.update(defaults)
            instance = self.__model(**kwargs)
            self.__storage[created_key] = instance
            created = True
        return instance, created

    def save(self, instance, arg_to_create: dict):
        try:
            self.__storage[self.__get_key(arg_to_create)] = instance
            return True
        except Exception as ex:
            print(ex)
            return False


class PKey:
    def __init__(self, fields: List['XlsField']):
        self.__fields = fields

    def from_row(self, ws: Worksheet, row: int):
        pkeys_values = []
        # print(f"{self.__fields = }")
        for field in self.__fields:
            value = field.get_value(ws, row)
            # print(f"pkey_field {value = }")
            if isinstance(value, dict):
                value = value.get("pkey")
            if field.verify(value, str(ws), row):
                # value is None and not field.is_blank():
                pkeys_values.append(value)
            else:
                _LOGGER.info(None, str(ws), f"Поле {field.get_db_name()}, PKey.from_row()")
                pkeys_values = []
                break
        # print(f"{pkeys_values = }")
        return self.__value(pkeys_values)

    def from_data(self, data: dict):
        pkeys_values = []
        for field in self.__fields:
            value = data.get(field.get_db_name())
            if isinstance(value, dict):
                value = value.get("pkey")
            if field.verify(value, None, None):
                pkeys_values.append(value)
            else:
                _LOGGER.info(None, None, f"Поле {field.get_db_name()}, PKey.from_data()")
                pkeys_values = []
                break
        return self.__value(pkeys_values)

    def get_order_num(self, ws: Worksheet, row: int):
        value = None
        try:
            value = ws.cell(row, 1).value
        except Exception:
            pass
        return value

    def __value(self, pkeys_values: list):
        result = None
        if any(pkeys_values):
            if len(pkeys_values) == 1:
                result = pkeys_values[0]
            else:
                result = tuple(pkeys_values)
        return result


class XlsModel:
    _pkey_field_name = "pkey"
    _row_field_name = "row"

    def __init__(self, name: str, ws: Worksheet,
                 first_data_row: int, fields: Dict[str, 'XlsField'],
                 extra_fields: Dict[str, 'XlsField'] = None,
                 ):
        self.__name = name
        self.__ws = ws
        self.__first_data_row = first_data_row
        self._add_extra_fields(fields, extra_fields)
        self.__fields = self.__get_updating_fields(fields)
        self.__pkey = self.__get_pkey_field()

    def is_data(self, xls_row: int = 0):
        if xls_row <= 0:
            return False
        if any(field.is_value(self.__ws, xls_row) for field in self.__fields.values()):
            return True
        else:
            return False

    def get(self, key=None, xls_row: int = 0):
        # print(f"{key = }  {xls_row = }")
        result = None
        if key is not None:
            xls_row = self.__get_row(key)
        if xls_row > 0:
            result = {}
            result[self._row_field_name] = xls_row

            # print(f"{self.__name}: {xls_row = }  {result['pkey'] = }")
            for field in self.__fields.values():
                # print(f"{self.__name = }  {field.get_db_name() =}  {xls_row = }")
                value = field.get_value(self.__ws, xls_row)
                # print(f"{self.__name = }  {field.get_db_name() =}  {xls_row = }  {value = }")
                if field.verify(value, str(self.__ws), xls_row):
                    result[field.get_db_name()] = value
                else:
                    _LOGGER.info(None, str(self.__ws), f"Поле {field.get_db_name()}, XlsModel.get()")
                    result = None
                    break
            if result and (pkey := self.__pkey.from_data(result)):
                result[self._pkey_field_name] = pkey
        return result

    def get_all_codes(self):
        """Возвращает множество всех значений поля с именем 'code'"""
        row = 2
        codes = set()
        while self.is_data(row):
            code = self.get_field("code", row)
            if isinstance(code, str):
                codes.add(code.strip())
            row += 1
        return codes

    def get_field(self, field_name, xls_row: int = 0):
        result = None
        if xls_row >= self.__first_data_row:
            field = self.__fields.get(field_name)
            if field and isinstance(self.__ws, Worksheet):
                result = field.get_value(self.__ws, xls_row)
        return result

    def get_name(self):
        return self.__name

    def get_pkey(self, entity: dict):
        """Получить значение ключа сущности"""
        return entity.get(self._pkey_field_name)

    def get_ws(self):
        return self.__ws

    def update_properties(cls, properties: dict, updated_properties: dict):
        """
        properties обновляется значениями updated_properties.
        Кроме служебных ключей.
        """
        for key, value in updated_properties.items():
            if key not in (cls._pkey_field_name, cls._row_field_name):
                properties[key] = value

    @classmethod
    def _add_extra_fields(cls, fields: Dict[str, 'XlsField'], extra_fields: Dict[str, 'XlsField']):
        """Расширяет поля множеством дополнительных полей"""
        if not isinstance(extra_fields, dict):
            return
        for name, field in extra_fields.items():
            if isinstance(field, XlsField):
                fields[name] = field

    def __get_list_ws(self, ws: List[Worksheet] | Worksheet):
        if isinstance(ws, Worksheet):
            return (ws,)
        elif isinstance(ws, Iterable):
            return tuple(wsheet for wsheet in ws if isinstance(wsheet, Worksheet))

    def __get_pkey_field(self):
        try:
            pkey_fields = []
            for field_key, field in self.__fields.items():
                if field.is_pkey():
                    pkey_fields.append(field)
            if len(pkey_fields):
                return PKey(pkey_fields)
            raise ValueError
        except Exception as ex:
            _LOGGER.error(self.__name, str(self.__ws), "Для XlsModel не удалось определить первичный ключ")
            raise ex

    def __get_row(self, key):
        result = 0
        # print(f"{key = }")
        # print(f"{self.__pkey.get_value(self.__ws, row) = }")
        # print(f"{self.__pkey.get_order_num(self.__ws, row) = }")
        row = self.__first_data_row
        while (self.is_data(row)):
            if self.__pkey.from_row(self.__ws, row) == key:
                result = row
                break
            row += 1
        return result

    def __get_updating_fields(self, fields: Dict[str, 'XlsField']):
        """Возвращаются поля с обновленными номерами колонок."""
        row = 1
        col = 1
        while (col_name := self.__ws.cell(row, col).value):
            col_name = str(col_name).strip()
            if field := fields.get(col_name):
                field.set_column_number(col)
            col += 1

        for col_name, field in fields.items():
            if field.get_column_number() is None:
                _LOGGER.warning(self.__name, str(self.__ws), f"Для поля '{col_name}' не найден столбец.")
            elif (fact_col_name := self.__ws.cell(row, field.get_column_number()).value) != col_name:
                if str(fact_col_name).strip() == col_name:
                    _LOGGER.warning(
                        self.__name, str(self.__ws),
                        f"Для поля '{col_name}' был указан номер столбца = {field.get_column_number()},"
                        f" на xls листе его название '{fact_col_name}' содержит пробелы в начале или конце."
                        f" Рекомендуется удалить эти пробелы.")
                else:
                    _LOGGER.error(
                        self.__name, str(self.__ws),
                        f"Для поля '{col_name}' был указан номер столбца = {field.get_column_number()},"
                        f" но на xls листе его название '{fact_col_name}'."
                        f" Для поля '{col_name}' номер столбца сброшен на None")
                    field.set_column_number(None)

        return fields


class XlsField:
    def __init__(self, db_name: str,
                 column_number: int = None,
                 p_key: bool = False, foreign_key: XlsModel = None,
                 blank: bool = False):
        self.__db_name = db_name
        self.__column = column_number
        self.__p_key = p_key
        self.__foreign_key = foreign_key
        self.__blank = blank

    def is_value(self, ws: Worksheet, row: int):
        try:
            if ws.cell(row, self.__column).value:
                return True
            else:
                return False
        except Exception:
            return False

    def get_value(self, ws: Worksheet, row: int):
        value = None
        try:
            value = ws.cell(row, self.__column).value
            if isinstance(value, str):
                value = value.strip()
        except Exception:
            pass
        # print(f"{self.__db_name = }  {row = }  {self.__column = }  {value = }")
        if self.verify(value, str(ws), row):
            # print(f"{value = } is not None")
            # print(f"{self.__foreign_key = }")
            if self.__foreign_key:
                result = self.__foreign_key.get(value)
            else:
                result = value
        else:
            _LOGGER.info(None, str(ws), f"Поле {self.get_db_name()}, XlsField.get_value()")
            result = None
        return result

    def get_order_num(self, ws: Worksheet, row: int):
        value = None
        try:
            value = ws.cell(row, 1).value
        except Exception:
            pass
        return value

    def is_blank(self):
        return self.__blank

    def is_pkey(self):
        return self.__p_key

    def is_foreign_key(self):
        return self.__foreign_key is not None

    def get_db_name(self):
        return self.__db_name

    def set_column_number(self, number):
        self.__column = number

    def get_column_number(self):
        return self.__column

    def verify(self, value, ws: str, row: int):
        """Проверка значения поля"""
        if not self.is_blank() and value is None:
            _LOGGER.error(None, ws,
                          f"Cтрока {row}, столбец {self.get_column_number()},"
                          f" {value = }. "
                          f" Поле '{self.get_db_name()}' не может быть пустым.")
            return False
        return True


class Result:
    """Результаты выполнения"""
    def __init__(self, action: str, type: str, logger: Logger):
        self.type = type
        self.logger = logger
        self.action = action
        self.available = []
        self.existing = []
        self.created = []
        self.not_created = []
        self.updated = []
        self.not_updated = []

    def write_result(self, created: bool, updated: bool, obj,
                     create_kwargs: dict = None, update_kwargs: dict = None,
                     changes_attrib: dict = None,
                     row=None):
        """
        Если 'created' == True, добавляет 'obj' в список 'create',
        иначе добавляет 'obj' в список 'no_create'. Если объект не None
        и 'created' == False, добавляет 'obj' в список 'exist'.
        Если объект не None
        и 'updated' == True, добавляет 'obj' в список 'update'
        """
        if not isinstance(create_kwargs, dict):
            create_kwargs = {}
        if not isinstance(update_kwargs, dict):
            update_kwargs = {}
        if not isinstance(changes_attrib, dict):
            changes_attrib = {}

        # если есть объект
        if obj:
            self.available.append(obj)
            # если объект создан
            if created:
                self.created.append(obj)
                self.logger.operation_info(
                    "success",
                    result_create_str(self.type, True,
                                      pk=getattr(obj, "pk"), row=row,
                                      **{**create_kwargs, **update_kwargs})
                )
            else:
                filter_f = get_filter_func(are_attribs_equal, create_kwargs)
                if (not list(filter(filter_f, self.existing))
                        and not list(filter(filter_f, self.created))):
                    self.existing.append(obj)
                    self.logger.operation_info(
                        "existing",
                        instance_exists_str(self.type, pk=getattr(obj, "pk"),
                                            **{**create_kwargs, **update_kwargs})
                    )
            if update_kwargs:
                # если объект не создан И обновлен
                if not created and updated:
                    self.updated.append(obj)
                    self.logger.operation_info(
                        "updated",
                        result_update_str(self.type, True,
                                          obj=str(obj), row=row,
                                          pk=getattr(obj, "pk"),
                                          **changes_attrib))
                # если объект должен был быть обновлен НО он НЕ обновлен
                if updated is False:
                    self.not_updated.append(obj)
                    self.logger.operation_info(
                        "updated_fail",
                        result_update_str(self.type, False,
                                          obj=str(obj), row=row,
                                          pk=getattr(obj, "pk"), **update_kwargs))
        # если нет объекта
        else:
            self.not_created.append({**create_kwargs, **update_kwargs})
            self.logger.operation_info(
                "fail",
                result_create_str(self.type, False, row,
                                  **{**create_kwargs, **update_kwargs}))

    def result_to_log(self):
        """Записывает результат в лог"""
        self.logger.info(" -----  TOTAL  -----")
        self.logger.info(f"--- EXISTS '{self.type}' = {len(self.existing)}")
        self.logger.info(f"--- CREATED '{self.type}' = {len(self.created)}")
        self.logger.info(f"--- NOT CREATED '{self.type}' = {len(self.not_created)}")
        self.logger.info(f"--- UPDATED '{self.type}' = {len(self.updated)}")
        self.logger.info(f"--- NOT UPDATED '{self.type}' = {len(self.not_updated)}")
        self.end_message_to_log()

    def start_message_to_log(self):
        """Записывает сообщение о начале действий в лог"""
        self.logger.info(f" --- START {self.action} '{self.type}'! ---")

    def end_message_to_log(self):
        """Записывает сообщение об окончании действий в лог"""
        self.logger.info(f" --- END {self.action} '{self.type}'! ---")
        for suffix in ("", "success", "existing", "updated", "updated_fail", "fail"):
            self.logger.end_to_log(suffix, " -----------  END  -----------")
#   End class Result


class CellReader:
    """Читает значения ячеек в xls файле"""
    def __init__(self, ws: Worksheet, key_number_columns: Dict[str, int]):
        self.__ws = ws
        self.__key_number_columns = key_number_columns
        self.__row = 1
        self.__col = 1

    def set_row(self, row: int):
        self.__row = row

    def get_value(self, key_column: str):
        try:
            return self.__ws.cell(
                row=self.__row,
                column=self.__key_number_columns[key_column]
            ).value
        except Exception:
            return None


def get_filter_func(func, attribs):
    """Декоратор. Получить функцию фильтрации объектов по атрибутам attribs"""
    def result(*args, **kwargs):
        return func(*args, **kwargs, attribs=attribs)
    return result


def are_attribs_equal(entity, attribs: dict):
    """Проверяет на равенство значений данные аттрибуты сущности"""
    return all(getattr(entity, arg, None) == value for arg, value in attribs.items())


def result_create_str(type_obj: str, result, row, **kwargs):
    """Возвращает текстовое сообщение о результате создания объекта"""
    return f"Из строки: {row}. {type_obj.upper()} create {result}!  {kwargs}"


def result_update_str(type_obj: str, result, row, **kwargs):
    """Возвращает текстовое сообщение о результате обновления объекта"""
    return f"Из строки: {row}. {type_obj.upper()} update {result}!  {kwargs}"


def instance_exists_str(type_obj: str, **kwargs):
    """Возвращает текстовое сообщение о существовании объекта"""
    return f"{type_obj.upper()} exists! {kwargs}"


def str_to_number(str_in, type: str, elem: str):
    """"""
    mess = None
    result = None
    if isinstance(str_in, str):
        str_in = str_in.strip()
    types = {"int": int, "float": float}
    if (func := types.get(type)):
        try:
            result = func(str_in)
        except Exception:
            mess = f"Ошибка конвертации в {type}. '{elem}' = {str_in}"
    return result, mess


def to_str(value):
    if value and not isinstance(value, str):
        return str(value)
    return value


# Создание экземпляров Result для создаваемых объектов
result_acc_points = Result("create", AccessPoints._meta.verbose_name_plural, Logger(f"{log_path}create_access_points"))
result_schedules = Result("create", Schedules._meta.verbose_name_plural, Logger(f"{log_path}create_schedules"))
result_units = Result("create", MeasureUnits._meta.verbose_name_plural, Logger(f"{log_path}create_units"))
result_data_types = Result("create", DataTypes._meta.verbose_name_plural, Logger(f"{log_path}create_data_types"))
result_plot_types = Result("create", PlotTypes._meta.verbose_name_plural, Logger(f"{log_path}create_plot_types"))
result_signal_types = Result("create", SignalTypes._meta.verbose_name_plural, Logger(f"{log_path}create_signal_types"))
result_device_models = Result("create", DeviceModels._meta.verbose_name_plural, Logger(f"{log_path}create_device_models"))
result_device_types = Result("create", DeviceTypes._meta.verbose_name_plural, Logger(f"{log_path}create_device_types"))
result_devices = Result("create", Devices._meta.verbose_name_plural, Logger(f"{log_path}create_devices"))
result_assets = Result("create", Assets._meta.verbose_name_plural, Logger(f"{log_path}create_assets"))
result_asset_types = Result("create", AssetsType._meta.verbose_name_plural, Logger(f"{log_path}create_asset_types"))
result_sgn_category = Result("create", SignalСategories._meta.verbose_name_plural, Logger(f"{log_path}create_sgn_cat"))
result_sgn_groups = Result("create", SignalGroups._meta.verbose_name_plural, Logger(f"{log_path}create_sgn_group"))
result_sgn_guides = Result("create", SignalsGuide._meta.verbose_name_plural, Logger(f"{log_path}create_guides"))
result_eobjects = Result("create", Substations._meta.verbose_name_plural, Logger(f"{log_path}create_e_objects"))
result_modbus_types = Result("create", ModbusTypes._meta.verbose_name_plural, Logger(f"{log_path}create_modbus_types"))
result_signals = Result("create", Signals._meta.verbose_name_plural, Logger(f"{log_path}create_signals"))
result_protocol = Result("create", Protocols._meta.verbose_name_plural, Logger(f"{log_path}create_protocols"))
reslt_formula = Result("create", Formulas._meta.verbose_name_plural, Logger(f"{log_path}create_formulas"))
result_databus = Result("create", DatabusSources._meta.verbose_name_plural, Logger(f"{log_path}create_databus_sources"))
result_d_storages = Result("create", DynamicStorages._meta.verbose_name_plural, Logger(f"{log_path}create_dynamic_storages"))
result_chart_tabs = Result("create", ChartTabs._meta.verbose_name_plural, Logger(f"{log_path}create_chart_tabs"))
result_asset_type_chart_tabs = Result("create", AssetsTypeChartTabs._meta.verbose_name_plural, Logger(f"{log_path}create_asset_type_chart_tabs"))
result_sgn_tabs = Result("create", SignalsChartTabs._meta.verbose_name_plural, Logger(f"{log_path}create_sgn_tabs"))
result_geomap_elems = Result("create", GeoMap._meta.verbose_name_plural, Logger(f"{log_path}create_geo_map_elems"))


results = [result_acc_points, result_schedules, result_units, result_data_types,
           result_plot_types, result_signal_types, result_device_models,
           result_device_types, result_devices, result_assets, result_asset_types,
           result_sgn_category, result_sgn_groups, result_sgn_guides,
           result_modbus_types, result_signals, result_protocol,
           reslt_formula, result_databus, result_d_storages,
           result_chart_tabs, result_sgn_tabs,
           result_asset_type_chart_tabs, result_eobjects, result_geomap_elems]

print("start_message_to_log")
for result in results:
    result.start_message_to_log()

schedule = XlsModel(
    name="schedule",
    ws=d_models_wb["schedules"],
    first_data_row=2,
    fields={
        "name": XlsField("name", 2, p_key=True),
        "interval": XlsField("interval_seconds", 3)
    },
)

unit = XlsModel(
    name="unit",
    ws=d_models_wb["units"],
    first_data_row=2,
    fields={"name": XlsField("name", 3, p_key=True),
            "code": XlsField("code", 2)},
    )

databus_source = XlsModel(
    name="databus_source",
    ws=d_models_wb["signal_types"],
    first_data_row=2,
    fields={"code": XlsField("code", 2, p_key=True)},
    )

dynamic_storage = XlsModel(
    name="dynamic_storage",
    ws=d_models_wb["signal_types"],
    first_data_row=2,
    fields={"code": XlsField("code", 2, p_key=True)},
    )

data_type = XlsModel(
    name="data_type",
    ws=d_models_wb["data_types"],
    first_data_row=2,
    fields={"code": XlsField("code", 2),
            "name": XlsField("name", 3, p_key=True)},
    )

plot_type = XlsModel(
    name="plot_type",
    ws=d_models_wb["plot_types"],
    first_data_row=2,
    fields={
        "code": XlsField("code", 2),
        "name": XlsField("name", 3, p_key=True),
    },
)

signal_type = XlsModel(
    name="signal_type",
    ws=d_models_wb["signal_types"],
    first_data_row=2,
    fields={"code": XlsField("code", 2, ),
            "name": XlsField("name", 3, p_key=True)},
    )

signals_category = XlsModel(
    name="signal_category",
    ws=d_models_wb["signal_categories"],
    first_data_row=2,
    fields={"code": XlsField("code", 2, p_key=True),
            "name": XlsField("name", 3)},
    )

signals_group = XlsModel(
    name="signal_group",
    ws=d_models_wb["signal_groups"],
    first_data_row=2,
    fields={"code": XlsField("code", 2, p_key=True),
            "name": XlsField("name", 3)},
    )

signals_guide = XlsModel(
    name="signals_guide",
    ws=d_models_wb["signals_guide"],
    first_data_row=2,
    fields={
        "code": XlsField("code", None, p_key=True),
        "name": XlsField("name", None),
        "unit": XlsField("unit", None, foreign_key=unit, blank=True),
        "databus_source": XlsField("databus_source", None, foreign_key=databus_source, blank=True),
        "dynamic_storage": XlsField("dynamic_storage", None, foreign_key=dynamic_storage, blank=True),
        "signal_category": XlsField("category", None, foreign_key=signals_category, blank=True),
        "signal_group": XlsField("group", None, foreign_key=signals_group, blank=True),
        "signal_type": XlsField("sg_type", None, foreign_key=signal_type),
        "lim0_code": XlsField("lim0_code", None, blank=True),
        "lim1_code": XlsField("lim1_code", None, blank=True),
        "diag_code": XlsField("diag_code", None, blank=True),
        "in_plot": XlsField("in_plot", None, blank=True),
        "plot_type": XlsField("plot_type", None, foreign_key=plot_type, blank=True),
        "data_type": XlsField("data_type", None, foreign_key=data_type, blank=True),
        "mms_data_object": XlsField("mms_data_object", None, blank=True),
        "mms_logical_node": XlsField("mms_logical_node", None, blank=True),
        "mms_class": XlsField("mms_class", None, blank=True),
        "opc_label": XlsField("opc_label", None, blank=True),
        "relevance_span": XlsField("relevance_span", None, blank=True),
        "freeze_span": XlsField("freeze_span", None, blank=True),
        "precision": XlsField("precision", None, blank=True),
        "transformer": XlsField("app_to_transformer", None, blank=True),
        "bush": XlsField("app_to_bush", None, blank=True),
        "gis": XlsField("app_to_gis", None, blank=True),
        "breaker": XlsField("app_to_breaker", None, blank=True),
        "gil": XlsField("app_to_gil", None, blank=True),
        "ct": XlsField("app_to_ct", None, blank=True),
        "vt": XlsField("app_to_vt", None, blank=True),
        "cap_comm": XlsField("app_to_cap_comm", None, blank=True),
        "cap_bank": XlsField("app_to_cap_bank", None, blank=True),
        "surge_arrester": XlsField("app_to_surge_arrester", None, blank=True),
        "shunt_reactor": XlsField("app_to_shunt_reactor", None, blank=True),
        "switchgear": XlsField("app_to_switchgear", None, blank=True),
        "switchgear_assembly": XlsField("app_to_switchgear_assembly", None, blank=True),
        "disconnector": XlsField("app_to_disconnector", None, blank=True),
        "support_insulator": XlsField("app_to_support_insulator", None, blank=True),
        # "gases": XlsField("gases_tab", None, blank=True),
        # "humidity": XlsField("humidity_tab", None, blank=True),
        # "temperature": XlsField("temperature_tab", None, blank=True),
        # "power": XlsField("power_tab", None, blank=True),
        # "inputState": XlsField("inputState_tab", None, blank=True),
        # "wear": XlsField("wear_tab", None, blank=True),
        # "its": XlsField("its_tab", None, blank=True),
    },
)

signals_guide_update = XlsModel(
    name="signals_guide_update",
    ws=sgn_list_wb["signals_guide"],
    first_data_row=2,
    fields={
        "code": XlsField("code", None, p_key=True),
        "name": XlsField("name", None),
        # "databus_source": XlsField("databus_source", foreign_key=databus_source, blank=True),
        # "dynamic_storage": XlsField("dynamic_storage", foreign_key=dynamic_storage, blank=True),
        # "signal_category": XlsField("category", foreign_key=signals_category, blank=True),
        # "signal_group": XlsField("group", foreign_key=signals_group, blank=True),
        # "signal_type": XlsField("sg_type", foreign_key=signal_type),
    }
)

access_point = XlsModel(
    name="access_point",
    ws=sgn_list_wb["access_points"],
    first_data_row=2,
    fields={
        "code": XlsField("code", 2, p_key=True),
        "host": XlsField("ip", 4, blank=True),
        "port": XlsField("port", 5, blank=True),
        "url": XlsField("url", 6, blank=True),
        "com_port": XlsField("com_port", 7, blank=True),
        "baud_rate": XlsField("baud_rate", 8, blank=True),
        "data_bits": XlsField("data_bits", 9, blank=True),
        "stop_bits": XlsField("stop_bits", 10, blank=True),
        "parity": XlsField("parity", 11, blank=True),
        "flow_control": XlsField("flow_control", 12, blank=True),
    }
)

protocol = XlsModel(
    name="protocol",
    ws=d_models_wb["protocols"],
    first_data_row=2,
    fields={
        "code": XlsField("code", 2, p_key=True),
        "name": XlsField("name", 3, blank=True),
        "service": XlsField("service", 4, blank=True),
    }
)

device_type = XlsModel(
    name="device_type",
    ws=d_models_wb["device_types"],
    first_data_row=2,
    fields={
        "code": XlsField("code", 2, p_key=True),
        "name": XlsField("name", 3, blank=True),
    },
)

device_model = XlsModel(
    name="device_model",
    ws=d_models_wb["device_models"],
    first_data_row=2,
    fields={
        "code": XlsField("code", 2, p_key=True),
        "name": XlsField("name", 3, blank=True),
        "type": XlsField("device_type", 4, foreign_key=device_type, blank=True),
        "manufacturer": XlsField("manufacturer", 5, blank=True),
        "range": XlsField("measuring_range", 6, blank=True),
        "accuracy": XlsField("accuracy", 7, blank=True),
        "register_no": XlsField("register_no", 8, blank=True),
    },
)

device = XlsModel(
    name="device",
    ws=sgn_list_wb["devices"],
    first_data_row=2,
    fields={
        "code": XlsField("name", 3, p_key=True),
        "gateway": XlsField("access_point", 2, foreign_key=access_point, blank=True),
        "device_model": XlsField("model", 4, foreign_key=device_model, blank=True),
        "common_address": XlsField("common_address", 6, blank=True),
        "profile": XlsField("profile", 7, blank=True),
        "last_verification": XlsField("last_verification", 8, blank=True),
        "next_verification": XlsField("next_verification", 9, blank=True),
        "serial_number": XlsField("serial_number", 10, blank=True),
        "protocol": XlsField("protocol", 11, foreign_key=protocol, blank=True),
        "word_order": XlsField("wordorder", 12, blank=True),
        "byte_order": XlsField("byteorder", 13, blank=True),
        "mms_logical_device": XlsField("mms_logical_device", 14, blank=True),
    },
)

asset_type = XlsModel(
    name="asset_type",
    ws=d_models_wb["asset_types"],
    first_data_row=2,
    fields={
        "code": XlsField("code", 2, p_key=True),
        "name": XlsField("name", 3, blank=True),
    },
)

eobject = XlsModel(
    name="object",
    ws=sgn_list_wb["object"],
    first_data_row=2,
    fields={
        "code": XlsField("code", None, p_key=True),
        "name": XlsField("name", None),
        "type": XlsField("type", None),
        "parent_object": XlsField("parent", None, blank=True),
    },
)

geomap_elem = XlsModel(
    name="geomap_elem",
    ws=sgn_list_wb["object"],
    first_data_row=2,
    fields={
        "code": XlsField("code", None, p_key=True, foreign_key=eobject),
        "x_coordinate":  XlsField("x_coordinate", None, blank=True),
        "y_coordinate":  XlsField("y_coordinate", None, blank=True),
        "position":  XlsField("position", None, blank=True),
        "always_show":  XlsField("always_show", None, blank=True),
    },
)

asset = XlsModel(
    name="asset",
    ws=sgn_list_wb["assets"],
    first_data_row=2,
    fields={
        "code": XlsField("code", None, p_key=True),
        "name": XlsField("name", None),
        "model": XlsField("model", None, blank=True),
        "asset_type": XlsField("type", None, foreign_key=asset_type, blank=True),
        "mms_logical_device": XlsField("mms_logical_device", None, blank=True),
        "object": XlsField("substation", None, foreign_key=eobject, blank=True),
    },
)

modbus_type = XlsModel(
    name="modbus_type",
    ws=d_models_wb["modbus_types"],
    first_data_row=2,
    fields={
        "code": XlsField("code", 2, p_key=True),
        "name": XlsField("name", 3, blank=True),
    },
)

chart_tab = XlsModel(
    name="chart_tab",
    ws=ui_models_wb["tabs_guide"],
    first_data_row=2,
    fields={
        "code": XlsField("code", None, p_key=True),
        "name": XlsField("name", None, blank=True),
    },
)

asset_type_ch_tab = XlsModel(
    name="chart_tab",
    ws=ui_models_wb["tabs_asset_types"],
    first_data_row=2,
    fields={
        "code": XlsField("code", None, p_key=True),
        "asset_type": XlsField("asset_type", None, foreign_key=asset_type),
        "tab": XlsField("chart_tab", None, foreign_key=chart_tab),
    },
)

signal = XlsModel(
    name="signal",
    ws=sgn_list_wb["signals"],
    first_data_row=2,
    fields={
        "asset": XlsField("asset", None, foreign_key=asset, p_key=True),
        "device": XlsField("device", None, foreign_key=device, p_key=True),
        "code": XlsField("code", None, foreign_key=signals_guide, p_key=True),
        "unit_source": XlsField("unit_source", None, foreign_key=unit, blank=True),
        "processing_mask": XlsField("processing_mask", None, blank=True),
        "schedule": XlsField("schedule", None, foreign_key=schedule, blank=True),
        "address": XlsField("address", None, blank=True),
        "bit": XlsField("bit", None, blank=True),
        "modbus_function": XlsField("modbus_function", None, blank=True),
        "func_constr": XlsField("func_constr", None, blank=True),
        "value_path": XlsField("value_path", None, blank=True),
        "quality_path": XlsField("quality_path", None, blank=True),
        "timestamp_path": XlsField("timestamp_path", None, blank=True),
        "formula": XlsField("formula", None, blank=True),
        "value_type": XlsField("value_type", None, foreign_key=modbus_type, blank=True),
    },
)

signal_chart_tab = XlsModel(
    name="signal_chart_tab",
    ws=sgn_list_wb["signal_tabs"],
    first_data_row=2,
    fields={
        "no.":  XlsField("id", None, p_key=True),
        "code": XlsField("code", None, foreign_key=signals_guide),
        "tab": XlsField("chart_tab", None, foreign_key=asset_type_ch_tab),
    },
    extra_fields={
        code: XlsField(f"asset_{code}", None, blank=True)
        for code in asset.get_all_codes()
    },
)


class ScheduleManager:
    __db_model = Schedules
    __local_storage = LocalStorage(__db_model)
    __execution_result = result_schedules
    __manager = Manager(__db_model, __local_storage, __execution_result)

    @classmethod
    def get_instance(cls, args: dict, to_db: bool):
        if not args or not isinstance(args, dict):
            return None
        args = cls.__manager.check_is_dict(args)
        args_to_create = {
            "pkey": args.get("pkey"),
            "interval_seconds": args.get("interval_seconds")
        }
        args_to_update = {
            "name": args.get("name"),
        }
        return cls.__manager._get_instance(args_to_create, args_to_update,
                                           args.get("row"), to_db)


class UnitManager:
    __db_model = MeasureUnits
    __local_storage = LocalStorage(__db_model)
    __execution_result = result_units
    __manager = Manager(__db_model, __local_storage, __execution_result)

    @classmethod
    def get_instance(cls, args: dict, to_db: bool):
        if not args or not isinstance(args, dict):
            return None
        args = cls.__manager.check_is_dict(args)
        args_to_create = {
            "pkey": args.get("pkey"),
            "code": args.get("code"),
            "defaults": {"name": args.get("name")}
        }
        args_to_update = {"name": args.get("name")}
        return cls.__manager._get_instance(args_to_create, args_to_update,
                                           args.get("row"), to_db)


# Менеджеры для создания signal_guide
class DatabusSourceManager:
    __db_model = DatabusSources
    __local_storage = LocalStorage(__db_model)
    __execution_result = result_databus
    __manager = Manager(__db_model, __local_storage, __execution_result)

    @classmethod
    def get_instance(cls, args: dict, to_db: bool):
        if not args or not isinstance(args, dict):
            return None
        args = cls.__manager.check_is_dict(args)
        args_to_create = {
            "pkey": args.get("pkey"),
            "name": args.get("code")
        }
        args_to_update = {}
        return cls.__manager._get_instance(args_to_create, args_to_update,
                                           args.get("row"), to_db)


class DynamicStorageManager:
    __db_model = DynamicStorages
    __local_storage = LocalStorage(__db_model)
    __execution_result = result_d_storages
    __manager = Manager(__db_model, __local_storage, __execution_result)

    @classmethod
    def get_instance(cls, args: dict, to_db: bool):
        if not args or not isinstance(args, dict):
            return None
        args = cls.__manager.check_is_dict(args)
        args_to_create = {
            "pkey": args.get("pkey"),
            "name": args.get("code")
        }
        args_to_update = {}
        return cls.__manager._get_instance(args_to_create, args_to_update,
                                           args.get("row"), to_db)


class DataTypeManager:
    __db_model = DataTypes
    __local_storage = LocalStorage(__db_model)
    __execution_result = result_data_types
    __manager = Manager(__db_model, __local_storage, __execution_result)

    @classmethod
    def get_instance(cls, args: dict, to_db: bool):
        if not args or not isinstance(args, dict):
            return None
        args = cls.__manager.check_is_dict(args)
        args_to_create = {
            "pkey": args.get("pkey"),
            "code": args.get("code"),
        }
        args_to_update = {"name": args.get("name")}
        return cls.__manager._get_instance(args_to_create, args_to_update,
                                           args.get("row"), to_db)


class PlotTypeManager:
    __db_model = PlotTypes
    __local_storage = LocalStorage(__db_model)
    __execution_result = result_data_types
    __manager = Manager(__db_model, __local_storage, __execution_result)

    @classmethod
    def get_instance(cls, args: dict, to_db: bool):
        if not args or not isinstance(args, dict):
            return None
        args = cls.__manager.check_is_dict(args)
        args_to_create = {
            "pkey": args.get("pkey"),
            "code": args.get("code"),
        }
        args_to_update = {"name": args.get("name")}
        return cls.__manager._get_instance(args_to_create, args_to_update,
                                           args.get("row"), to_db)


class SignalTypeManager:
    __db_model = SignalTypes
    __local_storage = LocalStorage(__db_model)
    __execution_result = result_signal_types
    __manager = Manager(__db_model, __local_storage, __execution_result)

    @classmethod
    def get_instance(cls, args: dict, to_db: bool):
        if not args or not isinstance(args, dict):
            return None
        args = cls.__manager.check_is_dict(args)
        args_to_create = {
            "pkey": args.get("pkey"),
            "code": args.get("code"),
        }
        args_to_update = {"name": args.get("name")}
        return cls.__manager._get_instance(args_to_create, args_to_update,
                                           args.get("row"), to_db)


class SignalsCategoryManager:
    __db_model = SignalСategories
    __local_storage = LocalStorage(__db_model)
    __execution_result = result_sgn_category
    __manager = Manager(__db_model, __local_storage, __execution_result)

    @classmethod
    def get_instance(cls, args: dict, to_db: bool):
        if not args or not isinstance(args, dict):
            return None
        args = cls.__manager.check_is_dict(args)
        args_to_create = {
            "pkey": args.get("pkey"),
            "code": args.get("code")
        }
        args_to_update = {"name": args.get("name")}
        return cls.__manager._get_instance(args_to_create, args_to_update,
                                           args.get("row"), to_db)


# Менеджеры для создания прибора мониторинга
class AccessPointsManager:
    __db_model = AccessPoints
    __local_storage = LocalStorage(__db_model)
    __execution_result = result_acc_points
    __manager = Manager(__db_model, __local_storage, __execution_result)

    @classmethod
    def get_instance(cls, args: dict, to_db: bool):
        if not args or not isinstance(args, dict):
            return None

        args_to_create = {
            "pkey": args.get("pkey"),
            "code": args.get("code"),
        }
        args_to_update = {
            "ip": args.get("ip"),
            "port": args.get("port"),
            "url": args.get("url"),
            "com_port": args.get("com_port"),
            "baud_rate": args.get("baud_rate"),
            "data_bits": args.get("data_bits"),
            "stop_bits": args.get("stop_bits"),
            "parity": args.get("parity"),
            "flow_control": cls.format_flow_control(args.get("flow_control")),
        }
        instance = cls.__manager._get_instance(args_to_create, args_to_update,
                                               args.get("row"), to_db)
        return instance

    @staticmethod
    def format_flow_control(value):
        value = str(value).lower()
        if value == "yes":
            value = True
        elif value == "no":
            value = False
        else:
            value = None
        return value


class ProtocolsManager:
    __db_model = Protocols
    __local_storage = LocalStorage(__db_model)
    __execution_result = result_protocol
    __manager = Manager(__db_model, __local_storage, __execution_result)

    @classmethod
    def get_instance(cls, args: dict, to_db: bool):
        if not args or not isinstance(args, dict):
            return None
        args_to_create = {
            "pkey": args.get("pkey"),
            "code": args.get("code"),
        }
        args_to_update = {
            "name": args.get("name"),
            "listener": True if args.get("service").lower() == "listener" else False,
        }
        return cls.__manager._get_instance(args_to_create, args_to_update,
                                           args.get("row"), to_db)


class DeviceTypesManager:
    __db_model = DeviceTypes
    __local_storage = LocalStorage(__db_model)
    __execution_result = result_device_types
    __manager = Manager(__db_model, __local_storage, __execution_result)

    @classmethod
    def get_instance(cls, args: dict, to_db: bool):
        if not args or not isinstance(args, dict):
            return None
        args_to_create = {
            "pkey": args.get("pkey"),
            "code": args.get("code"),
        }
        args_to_update = {"name": args.get("name")}
        instance = cls.__manager._get_instance(args_to_create, args_to_update,
                                               args.get("row"), to_db)
        return instance


class DeviceModelsManager:
    __db_model = DeviceModels
    __local_storage = LocalStorage(__db_model)
    __execution_result = result_device_models
    __manager = Manager(__db_model, __local_storage, __execution_result)

    @classmethod
    def get_instance(cls, args: dict, to_db: bool):
        if not args or not isinstance(args, dict):
            return None
        args_to_create = {
            "pkey": args.get("pkey"),
            "code": args.get("code"),
        }
        args_to_update = {
            "name": args.get("name"),
            "device_type": DeviceTypesManager.get_instance(args.get("device_type"), to_db),
            "manufacturer": args.get("manufacturer"),
            "measuring_range": args.get("measuring_range"),
            "accuracy": args.get("accuracy"),
            "register_no": args.get("register_no"),
        }
        instance = cls.__manager._get_instance(args_to_create, args_to_update,
                                               args.get("row"), to_db)
        return instance


class DevicesManager:
    __db_model = Devices
    __local_storage = LocalStorage(__db_model)
    __execution_result = result_devices
    __manager = Manager(__db_model, __local_storage, __execution_result)

    @classmethod
    def get_instance(cls, args: dict, to_db: bool):
        if not args or not isinstance(args, dict):
            return None
        args_to_create = {
            "pkey": args.get("pkey"),
            "name": args.get("name"),
        }
        args_to_update = {
            "model": DeviceModelsManager.get_instance(args.get("model"), to_db),
            "access_point": AccessPointsManager.get_instance(args.get("access_point"), to_db),
            "common_address": args.get("common_address"),
            "wordorder": cls.format_order(args.get("wordorder")),
            "byteorder": cls.format_order(args.get("byteorder")),
            "protocol": ProtocolsManager.get_instance(args.get("protocol"), to_db),
            "mms_logical_device": args.get("mms_logical_device"),
        }
        instance = cls.__manager._get_instance(args_to_create, args_to_update,
                                               args.get("row"), to_db)
        return instance

    @staticmethod
    def format_order(value):
        value = str(value).lower()
        if value == "big":
            value = True
        elif value == "little":
            value = False
        else:
            value = None
        return value


# Менеджеры для создания актива
class AssetsTypeManager:
    __db_model = AssetsType
    __local_storage = LocalStorage(__db_model)
    __execution_result = result_asset_types
    __manager = Manager(__db_model, __local_storage, __execution_result)

    @classmethod
    def get_instance(cls, args: dict, to_db: bool):
        if not args or not isinstance(args, dict):
            return None
        args = cls.__manager.check_is_dict(args)
        args_to_create = {
            "pkey": args.get("pkey"),
            "code": args.get("code"),
            "defaults": {"name": args.get("name")}
        }
        args_to_update = {}
        return cls.__manager._get_instance(args_to_create, args_to_update,
                                           args.get("row"), to_db)


class EObjectsManager:
    __db_model = Substations
    __local_storage = LocalStorage(__db_model)
    __execution_result = result_eobjects
    __manager = Manager(__db_model, __local_storage, __execution_result)

    @classmethod
    def get_instance(cls, args: dict, to_db: bool) -> None | Substations:
        if not args or not isinstance(args, dict):
            return None
        args_to_create = {
            "pkey": args.get("pkey"),
            "name": args.get("name"),
            "type": cls._get_type(args.get("type")),
        }
        parent_args = eobject.get(args.get("parent"))
        args_to_update = {
            "parent": cls.get_instance(parent_args, to_db),
        }
        return cls.__manager._get_instance(args_to_create, args_to_update,
                                           args.get("row"), to_db)

    @classmethod
    def _get_type(cls, input_type: str):
        if not isinstance(input_type, str):
            input_type = ""
        input_type = input_type.lower().strip()
        if input_type == "организация":
            return "node"
        elif input_type == "объект":
            return "end_node"
        else:
            return None


class AssetsManager:
    __db_model = Assets
    __local_storage = LocalStorage(__db_model)
    __execution_result = result_assets
    __manager = Manager(__db_model, __local_storage, __execution_result)

    @classmethod
    def get_instance(cls, args: dict, to_db: bool):
        if not args or not isinstance(args, dict):
            return None
        args_to_create = {
            "pkey": args.get("pkey"),
            "name": args.get("name")
        }
        args_to_update = {
            "type": AssetsTypeManager.get_instance(args.get("type", {}), to_db),
            "model": args.get("model"),
            "mms_logical_device": args.get("mms_logical_device"),
            "substation": EObjectsManager.get_instance(args.get("substation", {}), to_db),
        }
        return cls.__manager._get_instance(args_to_create, args_to_update,
                                           args.get("row"), to_db)


# ОСНОВНОЙ менеджер для создания signal_guide
class SignalsGuideManager:
    __db_model = SignalsGuide
    __local_storage = LocalStorage(__db_model)
    __execution_result = result_sgn_guides
    __manager = Manager(__db_model, __local_storage, __execution_result)

    @classmethod
    def get_instance(cls, args: dict, to_db: bool):
        if not args or not isinstance(args, dict):
            return None
        args_to_create = {
            "pkey": args.get("pkey"),
            "code": args.get("code"),
            "defaults": {
                "name": args.get("name"),
                "sg_type": SignalTypeManager.get_instance(args.get("sg_type"), to_db),
            }
        }
        args_to_update = {
            "name": args.get("name"),
            "sg_type": SignalTypeManager.get_instance(args.get("sg_type"), to_db),
            "unit": UnitManager.get_instance(args.get("unit"), to_db),
            "category": SignalsCategoryManager.get_instance(args.get("category"), to_db),
            "group": SignalsGroupManager.get_instance(args.get("group"), to_db),
            "relevance_span": args.get("relevance_span"),
            "freeze_span": args.get("freeze_span"),
            "precision": args.get("precision"),
            "lim0_code": args.get("lim0_code"),
            "lim1_code": args.get("lim1_code"),
            "diag_code": args.get("diag_code"),
            "mms_data_object": args.get("mms_data_object"),
            "mms_logical_node": args.get("mms_logical_node"),
            "mms_class": args.get("mms_class"),
            "opc_label": args.get("opc_label"),
            "data_type": DataTypeManager.get_instance(args.get("data_type"), to_db),
            "in_plot": cls.in_plot_for_db(args.get("in_plot")),
            "plot_type": PlotTypeManager.get_instance(args.get("plot_type"), to_db),
            "databus_source": DatabusSourceManager.get_instance(args.get("databus_source"), to_db),
            "dynamic_storage": DynamicStorageManager.get_instance(args.get("dynamic_storage"), to_db),
        }
        instance = cls.__manager._get_instance(args_to_create, args_to_update,
                                               args.get("row"), to_db)
        return instance

    @classmethod
    def get_exist_instance(cls, args: dict, to_db: bool):
        if not args or not isinstance(args, dict):
            return None
        args_to_find = {
            "pkey": args.get("pkey"),
            "code": args.get("code"),
        }
        instance = cls.__manager._get_exist_instance(args_to_find, to_db)
        return instance

    @classmethod
    def in_plot_for_db(cls, value):
        if value is None:
            return None
        if not isinstance(value, str):
            value = str(value).lower()
        if value == "yes":
            return True
        elif value == "no":
            return False
        else:
            return None

    # @classmethod
    # def create_sgn_charttabs(cls, sgn_guide_instance: SignalsGuide, args: dict, to_db: bool):
    #     """Создает связку сигнала с вкладкой графиков"""
    #     tab_codes = [
    #         "gases", "humidity", "temperature", "power", "inputState", "wear", "its"]
    #     for code in tab_codes:
    #         is_tab: str = args.get(f"{code}_tab")
    #         if isinstance(is_tab, str) and is_tab.lower() == "yes":
    #             if code == "inputState":
    #                 code =="bushing"
    #             xls_chart_tab = chart_tab.get(key=code)
    #             if xls_chart_tab:
    #                 SignalsChartTabsManager.get_instance(
    #                     {
    #                         "row": args.get("row"),
    #                         "code": sgn_guide_instance,
    #                         "chart_tab": xls_chart_tab
    #                     },
    #                     to_db)


# Менеджеры для создания сигнала
class ModbusTypesManager:
    __db_model = ModbusTypes
    __local_storage = LocalStorage(__db_model)
    __execution_result = result_modbus_types
    __manager = Manager(__db_model, __local_storage, __execution_result)

    @classmethod
    def get_instance(cls, args: dict, to_db: bool):
        if not args or not isinstance(args, dict):
            return None
        args_to_create = {
            "pkey": args.get("pkey"),
            "code": args.get("code")
        }
        args_to_update = {"name": args.get("name")}
        return cls.__manager._get_instance(args_to_create, args_to_update,
                                           args.get("row"), to_db)


class FormulasManager:
    __db_model = Formulas
    __local_storage = LocalStorage(__db_model)
    __execution_result = reslt_formula
    __manager = Manager(__db_model, __local_storage, __execution_result)

    @classmethod
    def get_instance(cls, arg: str, to_db: bool):
        if not arg:
            return None
        args_to_create = {
            "pkey": arg,
            "name": arg,
            "expression": arg
        }
        args_to_update = {}
        return cls.__manager._get_instance(args_to_create, args_to_update,
                                           None, to_db)


# ОСНОВНОЙ менеджер для создания signal
class SignalsManager:
    __db_model = Signals
    __local_storage = LocalStorage(__db_model)
    __execution_result = result_signals
    __manager = Manager(__db_model, __local_storage, __execution_result)

    @classmethod
    def get_instance(cls, args: dict, to_db: bool):
        if not args or not isinstance(args, dict):
            return None
        code = SignalsGuideManager.get_instance(args.get("code"), to_db)
        asset = AssetsManager.get_instance(args.get("asset"), to_db)
        device = DevicesManager.get_instance(args.get("device"), to_db)
        # print(f"{code = }\n{asset = }\n{device = }")
        if not (code and asset and device):
            return None
        args_to_create = {
            "pkey": args.get("pkey"),
            "code": code,
            "asset": asset,
            "device": device,
        }
        args_to_update = {
            "address": args.get("address"),
            "bit": args.get("bit"),
            "value_type": ModbusTypesManager.get_instance(args.get("value_type"), to_db),
            "modbus_function": args.get("modbus_function"),
            "unit_source": UnitManager.get_instance(args.get("unit_source"), to_db),
            "func_constr": args.get("func_constr"),
            "value_path": args.get("value_path"),
            "quality_path": args.get("quality_path"),
            "timestamp_path": args.get("timestamp_path"),
            "schedule": ScheduleManager.get_instance(args.get("schedule"), to_db),
            "formula": FormulasManager.get_instance(args.get("formula"), to_db)
        }
        return cls.__manager._get_instance(args_to_create, args_to_update,
                                           args.get("row"), to_db)


# Менеджеры для настроек сигналов для фронта
class SignalsGroupManager:
    __db_model = SignalGroups
    __local_storage = LocalStorage(__db_model)
    __execution_result = result_sgn_groups
    __manager = Manager(__db_model, __local_storage, __execution_result)

    @classmethod
    def get_instance(cls, args: dict, to_db: bool):
        if not args or not isinstance(args, dict):
            return None
        args = cls.__manager.check_is_dict(args)
        args_to_create = {
            "pkey": args.get("pkey"),
            "code": args.get("code")
        }
        args_to_update = {"name": args.get("name")}
        return cls.__manager._get_instance(args_to_create, args_to_update,
                                           args.get("row"), to_db)


class ChartTabsManager:
    __db_model = ChartTabs
    __local_storage = LocalStorage(__db_model)
    __execution_result = result_chart_tabs
    __manager = Manager(__db_model, __local_storage, __execution_result)

    @classmethod
    def get_instance(cls, args: dict, to_db: bool):
        if not args or not isinstance(args, dict):
            return None
        args = cls.__manager.check_is_dict(args)
        args_to_create = {
            "pkey": args.get("pkey"),
            "code": args.get("code"),
        }
        args_to_update = {"name": args.get("name")}
        return cls.__manager._get_instance(args_to_create, args_to_update,
                                           args.get("row"), to_db)


class AssetsTypeChartTabsManager:
    __db_model = AssetsTypeChartTabs
    __local_storage = LocalStorage(__db_model)
    __execution_result = result_asset_type_chart_tabs
    __manager = Manager(__db_model, __local_storage, __execution_result)

    @classmethod
    def get_instance(cls, args: dict, to_db: bool):
        if not args or not isinstance(args, dict):
            return None
        args = cls.__manager.check_is_dict(args)
        args_to_create = {
            "pkey": args.get("pkey"),
            "chart_tab": ChartTabsManager.get_instance(args.get("chart_tab", {}), to_db),
            "asset_type": AssetsTypeManager.get_instance(args.get("asset_type", {}), to_db)
        }
        args_to_update = {"code": args.get("code")}
        return cls.__manager._get_instance(args_to_create, args_to_update,
                                           args.get("row"), to_db)


class SignalsChartTabsManager:
    __db_model = SignalsChartTabs
    __local_storage = LocalStorage(__db_model)
    __execution_result = result_sgn_tabs
    __manager = Manager(__db_model, __local_storage, __execution_result)

    @classmethod
    def get_instance(cls, args: dict, to_db: bool):
        if not args or not isinstance(args, dict):
            return None
        _signal_guide = SignalsGuideManager.get_exist_instance(args.get("code"), to_db)
        _chart_tab = AssetsTypeChartTabsManager.get_instance(args.get("chart_tab", {}), to_db)
        if not _signal_guide or not _chart_tab:
            return None
        args_to_create = {
            "pkey": args.get("pkey"),
            "code": _signal_guide,
            "chart_tab": _chart_tab,
        }
        args_to_update = {}

        result = []
        prefix_asset_field = "asset_"
        for field, value in args.items():
            if prefix_asset_field not in field or value is None:
                continue
            if isinstance(value, str) and value.lower() == "yes":
                try:
                    asset_code = field.split(prefix_asset_field)[1]
                except Exception:
                    xls_asset = None
                else:
                    xls_asset = asset.get(key=asset_code)
                if xls_asset:
                    _asset = AssetsManager.get_instance(xls_asset, to_db)
                    if not _asset:
                        continue
                    args_to_create["asset"] = _asset
                    instance = cls.__manager._get_instance(args_to_create, args_to_update,
                                                           args.get("row"), to_db)
                    if instance:
                        result.append(instance)
        if result:
            return result
        else:
            return None


class GeoMapManager:
    __db_model = GeoMap
    __local_storage = LocalStorage(__db_model)
    __execution_result = result_geomap_elems
    __manager = Manager(__db_model, __local_storage, __execution_result)

    @classmethod
    def get_instance(cls, args: dict, to_db: bool):
        if not args or not isinstance(args, dict):
            return None
        args = cls.__manager.check_is_dict(args)
        _eobject = EObjectsManager.get_instance(args.get("code"), to_db)
        geometry = cls._get_point_geometry(args)
        properties = cls._get_point_properties(args, _eobject)
        if not geometry or not properties:
            return None
        args_to_create = {
            "pkey": args.get("pkey"),
            "collection_code": "Substations",
            "geometry": geometry,
            "properties": properties
        }
        args_to_update = {"linked_obj": _eobject, "description": _eobject.name}
        return cls.__manager._get_instance(args_to_create, args_to_update,
                                           args.get("row"), to_db)

    @classmethod
    def get_point_coords(cls, x: str | float, y: str | float):
        """Возвращает координаты точки (x, y)"""
        if isinstance(x, str):
            x = x.strip()
        if isinstance(y, str):
            y = y.strip()
        try:
            coord_x = float(x)
            coord_y = float(y)
        except Exception as ex:
            err_str = f"[ERROR] при получении координат {cls.__db_model}: {str(ex)}"
            cls.__execution_result.logger.info(err_str)
            coord_x = None
            coord_y = None
        return coord_x, coord_y

    @classmethod
    def _get_point_geometry(cls, args: dict):
        coord_x, coord_y = cls.get_point_coords(args.get("x_coordinate"), args.get("y_coordinate"))
        if coord_x and coord_y:
            return {
                "type": "Point",
                "coordinates": [coord_x, coord_y],
            }
        else:
            return None

    @classmethod
    def _get_point_properties(cls, args: dict, eobject: Substations):
        if eobject:
            return {
                "name": eobject.name,
                "position": args.get("position", "bottom"),
                "always_show": str(args.get("always_show", False)).lower()
            }
        else:
            return None


def is_processing_mask(row: int):
    processing_mask = signal.get_field("processing_mask", row)
    if isinstance(processing_mask, str) and processing_mask.strip().lower() == "yes":
        return True
    else:
        message = f"Строка = {row}. Нет признака создания сигнала."
        _LOGGER.info(signal.get_name(), signal.get_ws(), message)
        return False


def get_xls_instance(xls_entity: XlsModel, row):
    instance = xls_entity.get(xls_row=row)
    if not instance:
        message = (f"Строка = {row}. Словарь параметров пуст.")
        _LOGGER.error(xls_entity.get_name(), xls_entity.get_ws(), message)
        return None
    else:
        return instance


def update_sg_guide(sg_guide_props: dict):
    if isinstance(sg_guide_props, dict):
        updated_sg_guide = signals_guide_update.get(
            key=signals_guide.get_pkey(sg_guide_props))
    else:
        return {}
    if isinstance(updated_sg_guide, dict):
        signals_guide.update_properties(sg_guide_props, updated_sg_guide)
    return sg_guide_props


def is_updated_sguide(row: int, signal_entity: dict):
    sg_guide_props = update_sg_guide(signal.get_field("code", row))
    if not sg_guide_props:
        message = (f"Строка = {row}. Невозможно создать сигнал."
                   " Словарь параметров описания сигнала после обновления пуст.")
        _LOGGER.error(signal.get_name(), signal.get_ws(), message)
        return False
    else:
        signal_entity["code"] = sg_guide_props
        return True


def is_measuring_device(row: int, signal_entity: dict):
    device_code = str(signal_entity.get("device", {}).get("code")).strip().lower()
    if device_code in ("смт", "ручной ввод"):
        message = (f"Строка = {row}. Сигнал не создан,"
                   f" так как указан прибор '{device_code}'."
                   " Будет предпринята попытка создать описание сигнала.")
        _LOGGER.warning(signal.get_name(), signal.get_ws(), message)
        return False
    else:
        return True


def signal_processing(row):
    if not is_processing_mask(row):
        return
    if not (entity := get_xls_instance(signal, row)):
        return
    if not is_updated_sguide(row, entity):
        return
    if is_measuring_device(row, entity):
        instance = SignalsManager.get_instance(entity, to_db)
        if not instance:
            message = f"Строка = {row}. Не удалось получить сигнал."
            _LOGGER.error(signal.get_name(), signal.get_ws(), message)
    else:
        instance = SignalsGuideManager.get_instance(entity.get("code"), to_db)
        if not instance:
            message = f"Строка = {row}. Не удалось получить описание сигнала."
            _LOGGER.error(signal.get_name(), signal.get_ws(), message)


def signal_chart_tab_processing(row):
    if not (entity := get_xls_instance(signal_chart_tab, row)):
        return
    instance = SignalsChartTabsManager.get_instance(entity, to_db)
    if not instance:
        message = f"Строка = {row}. Не удалось получить соответствие (сигнал, вкладка графиков, оборудование)."
        _LOGGER.error(signal_chart_tab.get_name(), signal_chart_tab.get_ws(), message)


def geomap_processing(row):
    """Создает элементы карты"""
    if not (entity := get_xls_instance(geomap_elem, row)):
        return
    instance = GeoMapManager.get_instance(entity, to_db)
    if not instance:
        message = f"Строка = {row}. Не удалось получить точку энергообъекта для карты."
        _LOGGER.error(geomap_elem.get_name(), geomap_elem.get_ws(), message)


def update_map_center():
    """Обновляет координаты центра карты согласно координат энергообъектов"""
    x_sum = 0
    y_sum = 0
    count = 0
    for elem in GeoMap.objects.filter(collection_code="Substations"):
        coords = elem.geometry.get("coordinates")
        if isinstance(coords, (list, tuple)) and len(coords) > 1:
            x, y = GeoMapManager.get_point_coords(coords[0], coords[1])
            if x is not None and y is not None:
                x_sum += x
                y_sum += y
                count += 1
    if count:
        x_center = x_sum / count
        y_center = y_sum / count
        geo_sett = GeoMapSetting.objects.first()
        if not geo_sett:
            geo_sett = GeoMapSetting(default_zoom=5, min_zoom=4, max_zoom=12, rotation=0)
        geo_sett.center_x = x_center
        geo_sett.center_y = y_center
        geo_sett.save()


def update_assets_guid():
    """Обновляет GUID у активов, если он пуст."""
    for asset in Assets.objects.select_related("type", "substation").all():
        if not asset.guid:
           asset.guid = guid.generate()
           asset.save()


def main():
    """Основной метод для выполнения"""
    # Создание сигналов
    for xls_entity, description, func in (
            (signal, "СИГНАЛОВ", signal_processing),
            (signal_chart_tab, "СООТВЕТСТВИЯ (сигнал, вкладка графиков, оборудование)", signal_chart_tab_processing),
            (geomap_elem, "элементов карты", geomap_processing),
    ):

        row = 2
        print(f"* * * * * * * * * * * * *  СОЗДАНИЕ {description} * * * * * * * * * * * * * *")
        print("row = ", row)
        while xls_entity.is_data(xls_row=row):
            func(row)
            row += 1
            print("* * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * *")
            print("row = ", row)

        print("result_to_log")
    update_map_center()

    # Логирование результатов
    for result in results:
        result.result_to_log()
    _LOGGER.info(None, None, " --- END! ---")


if __name__ == "__main__":
    main()
