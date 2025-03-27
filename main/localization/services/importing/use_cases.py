from typing import List
from pathlib import Path
from openpyxl import Workbook, load_workbook
from .manager import ImportManager
from .import_schemes import (imports_for_signal_localization,
                             imports_all_data)


def __import_by_manager_list(wb: Workbook, manager_list: List[ImportManager]):
    """Импортировать по списку менеджеров."""
    cache = {}
    for manager in manager_list:
        manager.set_cache(cache)
        manager.import_all(wb)

def delete_all(manager_list: List[ImportManager]):
    """Удаляет из базы данных все записи для данных менеджеров."""
    for manager in manager_list[::-1]:
        manager.delete_all()

def import_in_migration(localization_file: Path, import_mngs: list[ImportManager]):
    """Миграция данных локализации во время начальных миграций"""
    __import_by_manager_list(
        wb=load_workbook(localization_file, data_only=True),
        manager_list=import_mngs
    )


def import_all_data(localization_file: Path, clear_db: bool = False):
    """Импорт всех данных локализации из xls"""
    if clear_db:
        delete_all(imports_all_data)
    __import_by_manager_list(
        wb=load_workbook(localization_file, data_only=True),
        manager_list=imports_all_data
    )


def import_signal_settings(localization_file: Path):
    """Импорт данных локализации сигналов"""
    __import_by_manager_list(
        wb=load_workbook(localization_file, data_only=True),
        manager_list=imports_for_signal_localization
    )
